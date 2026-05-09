from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from datetime import date
import models


# ── Styling helpers ────────────────────────────────────────────

def _thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill():
    return PatternFill("solid", fgColor="1F3864")   # dark navy


def _dept_fill():
    return PatternFill("solid", fgColor="D9E1F2")   # light blue


def _alt_fill():
    return PatternFill("solid", fgColor="F2F2F2")   # light grey


def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


# ── Column definitions ─────────────────────────────────────────
# Matches the exact source file column order (columns D onward, 0-based index mapped to Excel cols)

COLUMNS = [
    ("#",                                           5,   _center()),
    ("Empleado: Sociedad",                          32,  _left()),
    ("ID",                                          12,  _center()),
    ("Estado",                                      12,  _center()),
    ("Código proyecto",                             20,  _center()),
    ("Empleado",                                    30,  _left()),
    ("Empleado: ID",                                15,  _center()),
    ("Nomina del Mes",                              14,  _center()),
    ("Periodo:",                                    16,  _center()),
    ("Número de Días",                              14,  _center()),
    ("Horas Extras Diurnas",                        14,  _center()),
    ("Horas Extras Nocturnas",                      14,  _center()),
    ("Horas Extras Dominicales / Festivas",         14,  _center()),
    ("Horas Extras Dominicales / Festivas / Nocturnas", 14, _center()),
    ("Observaciones:\nFlujo de trabajo",            35,  _left(wrap=True)),
    ("Observaciones:\nFlujo de trabajo / Aprobaciones", 35, _left(wrap=True)),
    ("Creado:\nFecha",                              14,  _center()),
    ("Aprobado: Fecha",                             14,  _center()),
]

START_COL = 4   # column D (1-indexed)
START_DATA_ROW = 13  # row where column headers go (1-indexed), data starts at 14


def _write_col_headers(ws, row: int):
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    for i, (label, _, align) in enumerate(COLUMNS):
        col = START_COL + i
        cell = ws.cell(row=row, column=col)
        cell.value = label
        cell.font = font
        cell.fill = fill
        cell.alignment = _center(wrap=True)
        cell.border = _thin()
    ws.row_dimensions[row].height = 30


def _write_dept_header(ws, row: int, dept_name: str, num_cols: int):
    col = START_COL
    cell = ws.cell(row=row, column=col)
    cell.value = dept_name
    cell.font = Font(bold=True, name="Arial", size=10)
    cell.fill = _dept_fill()
    cell.alignment = _left()
    end_col = col + num_cols - 1
    if end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    ws.row_dimensions[row].height = 18


def _write_data_row(ws, row: int, seq: int, company_name: str, employee: models.Employee,
                     novelty: models.Novelty, period: models.PayPeriod, alt: bool):
    fill = _alt_fill() if alt else None
    base_font = Font(name="Arial", size=9)

    def w(col_offset, value, center=False):
        col = START_COL + col_offset
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = base_font
        if fill:
            cell.fill = fill
        cell.border = _thin()
        cell.alignment = _center() if center else _left(wrap=True)

    w(0, seq, center=True)
    w(1, company_name)
    w(2, employee.internal_code or "", center=True)
    w(3, "Aprobado" if novelty and novelty.status == models.NoveltyStatus.aprobado else "Pendiente", center=True)
    w(4, employee.project_code or "", center=True)
    w(5, employee.full_name)
    w(6, employee.cedula, center=True)
    w(7, period.month_name, center=True)
    w(8, period.period_label, center=True)
    w(9, novelty.num_dias if novelty else 15.0, center=True)
    w(10, novelty.horas_extras_diurnas if novelty and novelty.horas_extras_diurnas else None, center=True)
    w(11, novelty.horas_extras_nocturnas if novelty and novelty.horas_extras_nocturnas else None, center=True)
    w(12, novelty.horas_extras_dom_fest if novelty and novelty.horas_extras_dom_fest else None, center=True)
    w(13, novelty.horas_extras_dom_fest_noct if novelty and novelty.horas_extras_dom_fest_noct else None, center=True)
    w(14, novelty.observaciones if novelty else None)
    w(15, novelty.observaciones_aprobacion if novelty else None)

    created_str = novelty.created_at.strftime("%Y-%m-%d") if novelty and novelty.created_at else ""
    approved_str = novelty.approved_at.strftime("%Y-%m-%d") if novelty and novelty.approved_at else ""
    w(16, created_str, center=True)
    w(17, approved_str, center=True)

    ws.row_dimensions[row].height = 20


def _write_company_header(ws, company: models.Company, period: models.PayPeriod, total_employees: int):
    """Writes the top metadata block for a company sheet."""
    today = date.today().strftime("%d de %B de %Y").lower()
    # capitalize month name
    parts = today.split()
    parts[3] = parts[3].capitalize()
    today = " ".join(parts)

    period_range = f"01-{period.month:02d}-{period.year} - 15-{period.month:02d}-{period.year}" \
        if period.period_number == 1 \
        else f"16-{period.month:02d}-{period.year} - {_days_in_month(period.month, period.year)}-{period.month:02d}-{period.year}"

    header_col = START_COL

    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 16

    def meta(row, label, value):
        c = ws.cell(row=row, column=header_col)
        c.value = label
        c.font = Font(bold=True, name="Arial", size=9)
        v = ws.cell(row=row, column=header_col + 1)
        v.value = value
        v.font = Font(name="Arial", size=9)

    ws.cell(row=3, column=header_col).value = (
        f"Administración 001  {company.full_name}  Nit. {company.nit or ''}"
    )
    ws.cell(row=3, column=header_col).font = Font(bold=True, name="Arial", size=11)

    ws.cell(row=4, column=header_col).value = (
        f"Fecha {today} Por Talento Humano Personal"
    )
    ws.cell(row=4, column=header_col).font = Font(name="Arial", size=9)

    ws.cell(row=5, column=header_col).value = "Flujo de trabajo: Buscar - Plantilla (Nomina Reporte)"
    ws.cell(row=5, column=header_col).font = Font(name="Arial", size=9)

    meta(7, "Creado", period_range)
    ws.cell(row=7, column=header_col + 4).value = "Categoría"
    ws.cell(row=7, column=header_col + 5).value = "20 — 02. TALENTO HUMANO"

    meta(8, "Tipo", "716 — 12.Reporte de Horas Laborada")
    ws.cell(row=8, column=header_col + 4).value = "Buscar"
    ws.cell(row=8, column=header_col + 5).value = "Requisiciones + Anexos"

    meta(9, "Estado", "Aprobado")
    meta(10, "Contar", "")
    ws.cell(row=11, column=header_col).value = total_employees
    ws.cell(row=11, column=header_col).font = Font(bold=True, name="Arial", size=10)


def _days_in_month(month: int, year: int) -> int:
    import calendar
    return calendar.monthrange(year, month)[1]


def _set_column_widths(ws):
    for i, (_, width, _) in enumerate(COLUMNS):
        col_letter = get_column_letter(START_COL + i)
        ws.column_dimensions[col_letter].width = width
    # freeze panes at the first data row
    ws.freeze_panes = f"{get_column_letter(START_COL)}14"


def generate_excel(period: models.PayPeriod, db: Session) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    companies = db.query(models.Company).filter(models.Company.is_active == True).all()

    for company in companies:
        departments = db.query(models.Department).filter(
            models.Department.company_id == company.id,
            models.Department.is_active == True,
        ).order_by(models.Department.name).all()

        # collect employees across all depts for this company in this period
        all_employees = db.query(models.Employee).filter(
            models.Employee.company_id == company.id,
            models.Employee.is_active == True,
        ).all()
        if not all_employees:
            continue

        novelty_map = {
            n.employee_id: n
            for n in db.query(models.Novelty).filter(
                models.Novelty.period_id == period.id,
                models.Novelty.employee_id.in_([e.id for e in all_employees]),
            ).all()
        }

        # Use sheet_name if set, otherwise short_name
        sheet_title = (company.sheet_name or company.short_name)[:31]
        ws = wb.create_sheet(title=sheet_title)

        _write_company_header(ws, company, period, len(all_employees))
        _set_column_widths(ws)

        current_row = START_DATA_ROW  # row 13: first col-headers block
        seq = 1
        alt = False

        for dept in departments:
            dept_employees = [e for e in all_employees if e.department_id == dept.id]
            if not dept_employees:
                continue

            # Column headers row
            _write_col_headers(ws, current_row)
            current_row += 1

            # Department name row
            _write_dept_header(ws, current_row, dept.name, len(COLUMNS))
            current_row += 1

            for emp in sorted(dept_employees, key=lambda e: e.full_name):
                novelty = novelty_map.get(emp.id)
                _write_data_row(ws, current_row, seq, company.full_name, emp, novelty, period, alt)
                seq += 1
                alt = not alt
                current_row += 1

            current_row += 1  # blank row between departments

        # Footer
        footer_cell = ws.cell(row=current_row, column=START_COL)
        footer_cell.value = f"TOTAL EMPLEADOS {company.full_name.upper()}: {len(all_employees)}"
        footer_cell.font = Font(bold=True, name="Arial", size=10)
        footer_cell.fill = PatternFill("solid", fgColor="1F3864")
        footer_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        end_col = START_COL + len(COLUMNS) - 1
        ws.merge_cells(
            start_row=current_row, start_column=START_COL,
            end_row=current_row, end_column=end_col,
        )

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
